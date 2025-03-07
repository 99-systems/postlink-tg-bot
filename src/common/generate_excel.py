import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.database.connection import db
from src.database.models import crud
from src.database.models.request import SendRequest, DeliveryRequest

def generate_excel_file():
    requests = crud.get_all_requests(db)

    if not requests:
        print("No data found in the database.")
        return None

    send_data = []
    delivery_data = []

    for request in requests:
        status = 'Открыта' if request.status == 'open' else 'Закрыта'
        user_info = f"{request.user.name}"
        if isinstance(request, SendRequest):
            send_data.append({
                "Пользователь": user_info,
                "Город отправления": request.from_location,
                "Город назначения": request.to_location,
                "Дата начала": request.from_date.strftime('%d.%m.%Y'),
                "Дата окончания": request.to_date.strftime('%d.%m.%Y'),
                "Размер груза": request.size_type,
                "Статус": status,
                "Описание": request.description,
                "Дата создания": request.created_at.strftime('%d.%m.%Y %H:%M'),
                "Дата изменения": request.updated_at.strftime('%d.%m.%Y %H:%M')
            })
        elif isinstance(request, DeliveryRequest):
            delivery_data.append({
                "Пользователь": user_info,
                "Город отправления": request.from_location,
                "Город назначения": request.to_location,
                "Дата доставки": request.delivery_date.strftime('%d.%m.%Y'),
                "Размер груза": request.size_type,
                "Статус": status,
                "Дата создания": request.created_at.strftime('%d.%m.%Y %H:%M'),
                "Дата изменения": request.updated_at.strftime('%d.%m.%Y %H:%M')
            })

    filenames = []
    
    if send_data:
        send_df = pd.DataFrame(send_data)
        send_filename = "Все заявки на отправление.xlsx"
        send_df.to_excel(send_filename, index=False)
        format_excel_file(send_filename)
        filenames.append(send_filename)

    if delivery_data:
        delivery_df = pd.DataFrame(delivery_data)
        delivery_filename = "Все заявки на доставку.xlsx"
        delivery_df.to_excel(delivery_filename, index=False)
        format_excel_file(delivery_filename)
        filenames.append(delivery_filename)

    return filenames if filenames else None

def format_excel_file(filename):
    """ Форматит дату в эксель файле """
    wb = load_workbook(filename)
    ws = wb.active

    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
        ws.row_dimensions[row[0].row].height = 20  

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 3  

    wb.save(filename)
